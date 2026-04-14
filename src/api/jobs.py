"""Async job queue for report generation.

Each report = 1 job. Agent runs in background worker.
Frontend polls for status. Classification checkpoint pauses the job.

Production: replace with Redis/Celery (Phase 5).
"""

from __future__ import annotations

import asyncio
import logging
import uuid
from dataclasses import dataclass, field
from enum import StrEnum
from typing import Any

logger = logging.getLogger(__name__)


class JobState(StrEnum):
    """Possible states for a report generation job."""

    PENDING = "pending"
    RUNNING = "running"
    CHECKPOINT = "checkpoint"  # Waiting for classification confirmation
    GENERATING = "generating"  # Report being generated
    COMPLETE = "complete"
    FAILED = "failed"


# Hebrew status messages for each phase
PHASE_MESSAGES: dict[str, str] = {
    "intake": "קליטת נתוני לקוח",
    "taba_analysis": 'ניתוח תב"עות חלות',
    "building_mapping": "מיפוי וסיווג מבנים",
    "classification_checkpoint": "ממתין לאישור סיווג מבנים",
    "usage_fees": "חישוב דמי שימוש",
    "permit_fees": "חישוב דמי היתר",
    "capitalization": "חישוב היוון",
    "split": "חישוב פיצול",
    "report_assembly": "הרכבת דוח",
    "review": "בקרה ואישור",
    "output": "הפקת פלט סופי",
    "complete": "הושלם",
    "failed": "נכשל",
}


@dataclass
class Job:
    """A single report generation job.

    Tracks the full lifecycle of a feasibility study from intake
    through report generation and output.
    """

    id: str
    state: JobState = JobState.PENDING
    phase: str = "intake"
    progress: int = 0
    intake_data: dict[str, Any] = field(default_factory=dict)
    uploaded_files: list[str] = field(default_factory=list)
    buildings: list[dict[str, Any]] = field(default_factory=list)
    result: dict[str, Any] | None = None
    error: str | None = None

    # Internal asyncio event for checkpoint synchronization
    _checkpoint_event: asyncio.Event = field(default_factory=asyncio.Event)
    _confirmed_buildings: list[dict[str, Any]] = field(default_factory=list)

    @property
    def message(self) -> str:
        """Get the Hebrew status message for the current phase."""
        if self.state == JobState.FAILED and self.error:
            return f"שגיאה: {self.error}"
        return PHASE_MESSAGES.get(self.phase, self.phase)


class JobQueue:
    """Async job queue with optional database persistence.

    Manages job lifecycle including submission, status polling,
    checkpoint pausing/resuming, and cleanup. Writes state transitions
    to PostgreSQL/SQLite when available.
    """

    def __init__(self) -> None:
        """Initialize the job queue."""
        self._jobs: dict[str, Job] = {}
        self._tasks: dict[str, asyncio.Task[None]] = {}
        self._db_available: bool = False

    async def init_db(self) -> None:
        """Initialize database connection and recover stuck jobs."""
        try:
            from config.database import create_tables, get_session_factory
            await create_tables()
            self._db_available = True
            logger.info("Database initialized for job persistence")
            await self._recover_checkpoint_jobs()
        except Exception as exc:
            logger.warning("Database not available, using in-memory only: %s", exc)
            self._db_available = False

    async def _recover_checkpoint_jobs(self) -> None:
        """Mark CHECKPOINT jobs as FAILED on restart (they can't be resumed)."""
        if not self._db_available:
            return
        try:
            from datetime import datetime, timezone
            from config.database import get_session_factory, jobs_table
            from sqlalchemy import update

            factory = await get_session_factory()
            async with factory() as session:
                stmt = (
                    update(jobs_table)
                    .where(jobs_table.c.state == "checkpoint")
                    .values(
                        state="failed",
                        error="שרת נכבה במהלך ממתין לאישור — יש להגיש מחדש",
                        updated_at=datetime.now(timezone.utc),
                    )
                )
                result = await session.execute(stmt)
                await session.commit()
                if result.rowcount > 0:
                    logger.warning("Recovered %d stuck CHECKPOINT jobs → FAILED", result.rowcount)
        except Exception as exc:
            logger.error("Checkpoint recovery failed: %s", exc)

    async def _persist_job(self, job: Job) -> None:
        """Write job state to database (fire-and-forget)."""
        if not self._db_available:
            return
        try:
            from datetime import datetime, timezone
            from config.database import get_engine, get_session_factory, jobs_table

            now = datetime.now(timezone.utc)
            engine = await get_engine()
            dialect_name = engine.dialect.name  # "sqlite" or "postgresql"

            factory = await get_session_factory()
            async with factory() as session:
                values = {
                    "id": job.id,
                    "state": job.state,
                    "phase": job.phase,
                    "progress": job.progress,
                    "owner_name": job.intake_data.get("owner_name", ""),
                    "moshav_name": job.intake_data.get("moshav_name", ""),
                    "gush": job.intake_data.get("gush", 0),
                    "helka": job.intake_data.get("helka", 0),
                    "intake_data": job.intake_data,
                    "buildings": job.buildings or None,
                    "result": job.result,
                    "error": job.error,
                    "created_at": now,
                    "updated_at": now,
                }
                update_set = {
                    "state": values["state"],
                    "phase": values["phase"],
                    "progress": values["progress"],
                    "buildings": values["buildings"],
                    "result": values["result"],
                    "error": values["error"],
                    "updated_at": now,
                }

                if dialect_name == "postgresql":
                    from sqlalchemy.dialects.postgresql import insert as pg_insert
                    stmt = pg_insert(jobs_table).values(**values)
                    stmt = stmt.on_conflict_do_update(index_elements=["id"], set_=update_set)
                else:
                    from sqlalchemy.dialects.sqlite import insert as sqlite_insert
                    stmt = sqlite_insert(jobs_table).values(**values)
                    stmt = stmt.on_conflict_do_update(index_elements=["id"], set_=update_set)

                await session.execute(stmt)
                await session.commit()
        except Exception as exc:
            logger.warning("Failed to persist job %s to DB: %s", job.id, exc)

    async def submit(self, intake: dict[str, Any]) -> str:
        """Submit a new job.

        Args:
            intake: Intake form data matching Nachla model fields.

        Returns:
            Generated job_id string.
        """
        job_id = str(uuid.uuid4())
        job = Job(id=job_id, intake_data=intake)
        self._jobs[job_id] = job

        logger.info("Job %s submitted for %s", job_id, intake.get("owner_name", "unknown"))
        await self._persist_job(job)

        # Start background processing
        task = asyncio.create_task(self._run_job(job_id))
        self._tasks[job_id] = task

        return job_id

    async def get_status(self, job_id: str) -> Job | None:
        """Get current job status.

        Args:
            job_id: Job identifier.

        Returns:
            Job object if found, None otherwise.
        """
        return self._jobs.get(job_id)

    async def add_files(self, job_id: str, file_names: list[str]) -> None:
        """Add uploaded file references to a job.

        Args:
            job_id: Job identifier.
            file_names: List of uploaded file names.
        """
        job = self._jobs.get(job_id)
        if job:
            job.uploaded_files.extend(file_names)
            logger.info("Job %s: added %d files", job_id, len(file_names))

    async def pause_for_checkpoint(self, job_id: str, buildings: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Pause job at classification checkpoint and wait for user confirmation.

        CRITICAL: This implements the mandatory checkpoint from workflow step 3.4.
        The job will not proceed until the user explicitly confirms classifications.

        Args:
            job_id: Job identifier.
            buildings: List of building dicts to present for classification.

        Returns:
            The user-confirmed building list (may be modified by user).
        """
        job = self._jobs.get(job_id)
        if not job:
            raise ValueError(f"Job {job_id} not found")

        job.state = JobState.CHECKPOINT
        job.phase = "classification_checkpoint"
        job.buildings = buildings
        job._checkpoint_event.clear()

        logger.info("Job %s paused at classification checkpoint with %d buildings", job_id, len(buildings))
        await self._persist_job(job)

        # Wait for user confirmation (blocks until resume_after_checkpoint is called)
        await job._checkpoint_event.wait()

        logger.info("Job %s resumed after checkpoint", job_id)
        return job._confirmed_buildings

    async def resume_after_checkpoint(self, job_id: str, confirmed_buildings: list[dict[str, Any]]) -> None:
        """Resume job after user confirms classifications.

        Args:
            job_id: Job identifier.
            confirmed_buildings: User-confirmed building classifications.
        """
        job = self._jobs.get(job_id)
        if not job:
            raise ValueError(f"Job {job_id} not found")

        if job.state != JobState.CHECKPOINT:
            raise ValueError(f"Job {job_id} is not at checkpoint (state: {job.state})")

        job._confirmed_buildings = confirmed_buildings
        job.buildings = confirmed_buildings
        job.state = JobState.RUNNING
        job.phase = "usage_fees"

        # Signal the waiting coroutine to continue
        job._checkpoint_event.set()

        logger.info("Job %s checkpoint confirmed with %d buildings", job_id, len(confirmed_buildings))

    async def _run_job(self, job_id: str) -> None:
        """Execute the job in background using the real NachlaAgent.

        Orchestrates the full workflow: document analysis, building mapping,
        classification checkpoint, calculations, report generation, and output.

        Args:
            job_id: Job identifier.
        """
        from agent.main_agent import NachlaAgent
        from agent.system_prompt import build_system_prompt
        from config.settings import get_settings
        from models.building import Building
        from models.nachla import Nachla

        job = self._jobs.get(job_id)
        if not job:
            return

        try:
            job.state = JobState.RUNNING
            settings = get_settings()

            # Initialize Monday.com client (fire-and-forget, never blocks)
            monday_client = None
            monday_item_id = job.intake_data.get("monday_item_id")
            if monday_item_id:
                from integrations.monday_client import MondayClient
                monday_client = MondayClient()
                if monday_client.is_configured:
                    asyncio.create_task(monday_client.update_status(monday_item_id, "בבדיקה"))
                else:
                    monday_client = None

            # --- Create agent ---
            agent = NachlaAgent(settings)

            # Attach LLM client if API key is configured
            llm_client = None
            if settings.anthropic_api_key:
                from agent.llm_client import LLMClient
                llm_client = LLMClient(settings, audit_logger=agent.audit_logger)
                agent.attach_llm(llm_client)
                logger.info("Job %s: LLM client attached", job_id)
            else:
                logger.warning("Job %s: No ANTHROPIC_API_KEY, running without AI", job_id)

            # --- Build Nachla model from intake data ---
            nachla = Nachla(**job.intake_data)

            # Pre-populate tabas from user input (entered on Taba page before submission)
            pre_tabas = job.intake_data.get("pre_extracted_tabas")
            logger.info("Job %s: pre_extracted_tabas=%s count=%d",
                       job_id, pre_tabas is not None, len(pre_tabas) if pre_tabas else 0)
            if pre_tabas:
                from models.taba import Taba, TabaRights
                for td in pre_tabas:
                    try:
                        taba_dict = dict(td)
                        # Convert flat main/service area to unit_rights list
                        if "main_area_sqm" in taba_dict and "unit_rights" not in taba_dict:
                            main = float(taba_dict.pop("main_area_sqm", 0))
                            service = float(taba_dict.pop("service_area_sqm", 0))
                            num_units = int(float(taba_dict.get("num_units_allowed", 2)))
                            taba_dict["unit_rights"] = [
                                {"main_area_sqm": main, "service_area_sqm": service}
                            ] * max(1, num_units)
                        taba_dict.pop("source", None)
                        nachla.tabas.append(Taba(**taba_dict))
                    except Exception as exc:
                        logger.error("Failed to parse pre-extracted taba: %s", exc, exc_info=True)
                logger.info("Job %s: Pre-loaded %d tabas from user input", job_id, len(nachla.tabas))

            # Build system prompt
            priority = nachla.priority_area.value if nachla.priority_area else None
            agent.system_prompt = build_system_prompt(priority_area=priority)

            # --- Phase 1: Intake ---
            job.phase = "intake"
            job.progress = 5
            await agent._run_intake(nachla)

            # --- Phase 2: Taba analysis ---
            job.phase = "taba_analysis"
            job.progress = 15
            tabas = await agent._run_taba_analysis(nachla)
            if monday_client and monday_item_id:
                asyncio.create_task(monday_client.update_status(monday_item_id, 'ניתוח תב"ע הושלם'))

            # --- Phase 3: Building mapping ---
            job.phase = "building_mapping"
            job.progress = 30
            from agent.workflow import WorkflowPhase

            # Check if buildings were pre-extracted (user already validated at upload time)
            pre_buildings = job.intake_data.get("pre_extracted_buildings")
            pre_tabas = job.intake_data.get("pre_extracted_tabas")

            if pre_buildings:
                # Use pre-extracted buildings — skip LLM analysis
                logger.info("Job %s: Using %d pre-extracted buildings", job_id, len(pre_buildings))
                buildings = []
                for bd in pre_buildings:
                    try:
                        buildings.append(Building(**bd))
                    except Exception as exc:
                        logger.warning("Failed to parse pre-extracted building: %s", exc)
                nachla.buildings = buildings
                # Tabas already loaded from pre_extracted_tabas before Phase 1
            else:
                # No pre-extracted data — run LLM document analysis
                # Wait briefly for file uploads to arrive (uploaded async after job creation)
                if not job.uploaded_files:
                    for _ in range(10):
                        await asyncio.sleep(1)
                        if job.uploaded_files:
                            break

                uploaded_files: dict[str, str] = {}
                for fpath in job.uploaded_files:
                    if "survey" in fpath.lower() or "מדידה" in fpath:
                        uploaded_files["survey_map"] = fpath
                    elif "permit" in fpath.lower() or "היתר" in fpath:
                        uploaded_files["building_permits"] = fpath
                    else:
                        uploaded_files[fpath] = fpath
                logger.info("Job %s: %d uploaded files available", job_id, len(uploaded_files))
                buildings = await agent._run_building_mapping(nachla, uploaded_files)

            if monday_client and monday_item_id:
                asyncio.create_task(monday_client.update_status(monday_item_id, "מיפוי מבנים הושלם"))

            # --- Phase 3.4: Classification checkpoint ---
            if buildings and not pre_buildings:
                # Only pause for checkpoint if buildings came from LLM (not pre-validated)
                job.buildings = [b.model_dump() for b in buildings]
                confirmed_dicts = await self.pause_for_checkpoint(job_id, job.buildings)
                buildings = []
                for bd in confirmed_dicts:
                    try:
                        buildings.append(Building(**bd))
                    except Exception as exc:
                        logger.warning("Failed to parse confirmed building: %s", exc)

            # Mark phases as completed
            agent.workflow.confirm_classifications()
            for phase in (WorkflowPhase.BUILDING_MAPPING, WorkflowPhase.CLASSIFICATION, WorkflowPhase.CHECKPOINT):
                if phase not in agent.workflow.completed_phases:
                    agent.workflow.completed_phases.append(phase)

            if not buildings:
                logger.warning("Job %s: No buildings to process", job_id)

            # --- Phases 4-11: Calculations ---
            job.phase = "calculations"
            job.progress = 50
            calc_results = await agent._run_calculations(nachla, buildings, tabas)

            # --- Phase 12: Report assembly + narratives ---
            job.phase = "report"
            job.progress = 80
            job.state = JobState.GENERATING
            if monday_client and monday_item_id:
                asyncio.create_task(monday_client.update_status(monday_item_id, "טיוטה מוכנה"))
            report_data = await agent._build_report_data(nachla, buildings, tabas, calc_results)

            # --- Phase 13: Review (sanity checks) ---
            job.phase = "review"
            job.progress = 90
            await agent._run_review(report_data)

            # --- Phase 14: Output (generate documents) ---
            job.phase = "output"
            job.progress = 95

            output_dir = settings.output_directory
            word_path = None
            audit_path = None

            try:
                from documents.report_builder import build_report
                from pathlib import Path

                Path(output_dir).mkdir(parents=True, exist_ok=True)
                word_path = build_report(
                    nachla=nachla.model_dump(),
                    buildings=[b.model_dump() for b in buildings],
                    tabas=[t.model_dump() for t in tabas],
                    calc_results=calc_results,
                    report_date=report_data.report_date,
                    output_path=f"{output_dir}/{job_id}.docx",
                )
                logger.info("Job %s: Word report generated at %s", job_id, word_path)
            except Exception as exc:
                logger.error("Job %s: Word generation failed: %s", job_id, exc, exc_info=True)

            # Generate Excel summary
            excel_path = None
            try:
                from pathlib import Path as P
                P(output_dir).mkdir(parents=True, exist_ok=True)
                excel_path = f"{output_dir}/{job_id}.xlsx"
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "סיכום תחשיבים"
                ws.sheet_view.rightToLeft = True

                # Header
                ws.append(["סיכום בדיקת התכנות נחלה"])
                ws.append([f"בעל נחלה: {nachla.owner_name}", f"מושב: {nachla.moshav_name}"])
                ws.append([f"גוש: {nachla.gush}", f"חלקה: {nachla.helka}"])
                ws.append([])

                # Cost summary
                ws.append(["סוג תשלום", "סכום (ש\"ח)"])
                ws.append(["דמי שימוש", report_data.total_usage_fees])
                ws.append(["דמי היתר", report_data.total_permit_fees])
                h375 = calc_results.get("hivun", {}).get("hivun_375", {}).get("result", 0)
                h33 = calc_results.get("hivun", {}).get("hivun_33", {}).get("result", 0)
                ws.append(["היוון 3.75%", h375])
                ws.append(["היוון 33%", h33])
                ws.append([])

                # Buildings
                ws.append(["מבנים"])
                ws.append(["#", "שם", "סוג", "סטטוס", "שטח עיקרי", "חריגה"])
                for b in buildings:
                    type_he = {'residential': 'בית מגורים', 'service': 'שירות', 'agricultural': 'חקלאי',
                               'pool': 'בריכה', 'pergola': 'פרגולה', 'plach': 'פל"ח', 'shed_open': 'סככה'}
                    status_he = {'compliant': 'תקין', 'deviation': 'חריגה', 'no_permit': 'ללא היתר',
                                 'marked_demolition': 'להריסה'}
                    ws.append([
                        b.id, b.name,
                        type_he.get(b.building_type.value, b.building_type.value),
                        status_he.get(b.status.value, b.status.value),
                        b.main_area_sqm, b.deviation_sqm or 0,
                    ])
                ws.append([])

                # Tabas
                if tabas:
                    ws.append(["תב\"עות"])
                    ws.append(["מספר", "שם", "סטטוס", "שטח מגרש", "יח' דיור"])
                    for t in tabas:
                        ws.append([t.taba_number, t.taba_name, t.status, t.plot_size_sqm, t.num_units_allowed])

                # Style header
                from openpyxl.styles import Font
                ws['A1'].font = Font(bold=True, size=14)
                for cell in ws[5]:
                    if cell.value:
                        cell.font = Font(bold=True)

                wb.save(excel_path)
                logger.info("Job %s: Excel report generated at %s", job_id, excel_path)
            except Exception as exc:
                logger.error("Job %s: Excel generation failed: %s", job_id, exc)
                excel_path = None

            # Save audit log
            try:
                from pathlib import Path as P
                P(output_dir).mkdir(parents=True, exist_ok=True)
                audit_path = f"{output_dir}/{job_id}_audit.json"
                agent.save_audit_log(audit_path)
                logger.info("Job %s: Audit log saved at %s", job_id, audit_path)
            except Exception as exc:
                logger.error("Job %s: Audit log save failed: %s", job_id, exc)

            # --- Complete ---
            job.phase = "complete"
            job.progress = 100
            job.state = JobState.COMPLETE
            job.result = {
                "word_path": word_path,
                "excel_path": excel_path,
                "audit_path": audit_path,
                "pdf_path": None,
                "total_regularization_cost": report_data.total_regularization_cost,
                "total_usage_fees": report_data.total_usage_fees,
                "total_permit_fees": report_data.total_permit_fees,
                "betterment_levy": report_data.betterment_levy if hasattr(report_data, "betterment_levy") else 0,
                "hivun_375_total": (
                    (report_data.hivun_375_result.get("result") or report_data.hivun_375_result.get("total_cost", 0))
                    if isinstance(report_data.hivun_375_result, dict)
                    else 0
                ),
                "hivun_33_total": (
                    (report_data.hivun_33_result.get("result") or report_data.hivun_33_result.get("total_cost", 0))
                    if isinstance(report_data.hivun_33_result, dict)
                    else 0
                ),
                "token_usage": llm_client.token_usage if llm_client else {},
                "cost_estimate": llm_client.estimate_cost() if llm_client else {},
            }
            logger.info("Job %s: complete", job_id)
            await self._persist_job(job)
            if monday_client and monday_item_id:
                asyncio.create_task(monday_client.update_status(monday_item_id, "מאושר"))
                # Attach report files
                if word_path:
                    asyncio.create_task(monday_client.attach_file(monday_item_id, word_path))

        except asyncio.CancelledError:
            job.state = JobState.FAILED
            job.error = "העבודה בוטלה"
            logger.warning("Job %s cancelled", job_id)
            await self._persist_job(job)
        except Exception as exc:
            job.state = JobState.FAILED
            job.phase = "failed"
            job.error = str(exc)
            logger.exception("Job %s failed: %s", job_id, exc)
            await self._persist_job(job)
            if monday_client and monday_item_id:
                asyncio.create_task(monday_client.update_status(monday_item_id, "נכשל - דורש טיפול ידני"))

    async def list_jobs(self) -> list[dict[str, Any]]:
        """List all jobs, combining in-memory and DB sources.

        Returns:
            List of job summary dicts.
        """
        summaries: list[dict[str, Any]] = []
        seen_ids: set[str] = set()

        # In-memory jobs (current session)
        for job in self._jobs.values():
            summaries.append({
                "job_id": job.id,
                "owner_name": job.intake_data.get("owner_name", ""),
                "moshav_name": job.intake_data.get("moshav_name", ""),
                "status": job.state,
                "phase": job.phase,
                "created_at": None,
            })
            seen_ids.add(job.id)

        # DB jobs from previous sessions
        if self._db_available:
            try:
                from config.database import get_session_factory, jobs_table
                from sqlalchemy import select

                factory = await get_session_factory()
                async with factory() as session:
                    stmt = select(jobs_table).order_by(jobs_table.c.created_at.desc()).limit(100)
                    result = await session.execute(stmt)
                    for row in result:
                        if row.id not in seen_ids:
                            summaries.append({
                                "job_id": row.id,
                                "owner_name": row.owner_name,
                                "moshav_name": row.moshav_name,
                                "status": row.state,
                                "phase": row.phase,
                                "created_at": row.created_at.isoformat() if row.created_at else None,
                            })
            except Exception as exc:
                logger.warning("Failed to list DB jobs: %s", exc)

        return summaries

    async def cancel_job(self, job_id: str) -> bool:
        """Cancel a running job.

        Args:
            job_id: Job identifier.

        Returns:
            True if job was cancelled, False if not found or already complete.
        """
        task = self._tasks.get(job_id)
        if task and not task.done():
            task.cancel()
            return True
        return False

    async def shutdown(self) -> None:
        """Cancel all running jobs and clean up.

        Called during application shutdown.
        """
        for job_id, task in self._tasks.items():
            if not task.done():
                task.cancel()
                logger.info("Cancelled job %s during shutdown", job_id)

        # Wait for all tasks to complete cancellation
        if self._tasks:
            await asyncio.gather(*self._tasks.values(), return_exceptions=True)

        self._tasks.clear()
        logger.info("Job queue shutdown complete")

    def list_jobs_sync(self, status_filter: str | None = None) -> list[Job]:
        """List in-memory jobs synchronously, optionally filtered by status.

        For internal use only. The async list_jobs() method includes DB results.
        """
        jobs = list(self._jobs.values())
        if status_filter:
            jobs = [j for j in jobs if j.state == status_filter]
        return jobs

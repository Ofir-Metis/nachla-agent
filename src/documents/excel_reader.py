"""Read Excel reference tables for RMI lookups.

Uses openpyxl for direct cell access when the structure is known,
and pandas for generic table reading.

All string operations use UTF-8 encoding with ensure_ascii=False.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# Maximum file size for upload validation (50 MB).
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024

# Data older than this many days is considered stale.
STALE_THRESHOLD_DAYS = 90


class ExcelReader:
    """Read Excel reference tables for RMI lookups.

    Uses openpyxl for reading + pandas for data manipulation.
    """

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    @staticmethod
    def validate_file(file_path: str) -> None:
        """Validate the file exists, is an Excel file, and is under 50 MB."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        if not path.is_file():
            raise ValueError(f"Path is not a file: {file_path}")
        if path.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
            raise ValueError(f"File is not an Excel file (extension={path.suffix!r}): {file_path}")
        size = path.stat().st_size
        if size > MAX_FILE_SIZE_BYTES:
            raise ValueError(f"File exceeds 50 MB limit ({size / (1024 * 1024):.1f} MB): {file_path}")

    # ------------------------------------------------------------------
    # Specific table readers
    # ------------------------------------------------------------------

    def read_settlement_table(self, file_path: str) -> dict[str, float]:
        """Read settlement permit fee table.

        Returns:
            Mapping of settlement name to shovi value (value per sqm equivalent).
        """
        self.validate_file(file_path)
        rows = self._read_all_rows_openpyxl(file_path)
        if not rows:
            return {}

        # Auto-detect header row and find settlement name + numeric value columns.
        header_idx, headers = self._detect_header(rows)
        if header_idx is None:
            logger.warning("Could not detect header row in %s", file_path)
            return {}

        data_rows = rows[header_idx + 1 :]
        result: dict[str, float] = {}

        # Heuristic: first text column = settlement name, first numeric column = value.
        name_col, value_col = self._find_name_value_cols(headers, data_rows)
        if name_col is None or value_col is None:
            logger.warning("Could not identify name/value columns in %s", file_path)
            return {}

        for row in data_rows:
            if len(row) <= max(name_col, value_col):
                continue
            name = str(row[name_col]).strip() if row[name_col] is not None else ""
            raw_val = row[value_col]
            if not name:
                continue
            val = self._to_float(raw_val)
            if val is not None:
                result[name] = val

        logger.debug("Read %d settlement entries from %s", len(result), file_path)
        return result

    def read_plach_table(self, file_path: str) -> dict[str, float]:
        """Read PLACH rates table.

        Returns:
            Mapping of region or settlement name to rate.
        """
        self.validate_file(file_path)
        rows = self._read_all_rows_openpyxl(file_path)
        if not rows:
            return {}

        header_idx, headers = self._detect_header(rows)
        if header_idx is None:
            return {}

        data_rows = rows[header_idx + 1 :]
        name_col, value_col = self._find_name_value_cols(headers, data_rows)
        if name_col is None or value_col is None:
            return {}

        result: dict[str, float] = {}
        for row in data_rows:
            if len(row) <= max(name_col, value_col):
                continue
            name = str(row[name_col]).strip() if row[name_col] is not None else ""
            raw_val = row[value_col]
            if not name:
                continue
            val = self._to_float(raw_val)
            if val is not None:
                result[name] = val

        logger.debug("Read %d PLACH entries from %s", len(result), file_path)
        return result

    def read_rmi_decisions(self, file_path: str) -> list[dict]:
        """Read RMI decisions/guidelines table.

        Returns:
            List of dicts, one per row, using header names as keys.
        """
        return self.read_generic_table(file_path)

    def read_generic_table(
        self,
        file_path: str,
        sheet_name: str | None = None,
    ) -> list[dict]:
        """Read any Excel file into a list of row dicts using pandas.

        Args:
            file_path: Path to the Excel file.
            sheet_name: Optional sheet name. Uses the first sheet if None.

        Returns:
            List of dicts with column-name keys.
        """
        self.validate_file(file_path)
        try:
            import pandas as pd

            kwargs: dict = {"engine": "openpyxl"}
            if sheet_name is not None:
                kwargs["sheet_name"] = sheet_name

            df = pd.read_excel(file_path, **kwargs)
            # Drop fully empty rows
            df = df.dropna(how="all")
            # Convert to list of dicts, ensuring strings are UTF-8 clean
            records = df.to_dict(orient="records")
            logger.debug("Read %d rows from %s", len(records), file_path)
            return records
        except Exception:
            logger.warning("Failed to read generic table from %s", file_path, exc_info=True)
            return []

    def get_table_metadata(self, file_path: str) -> dict:
        """Get metadata about an Excel file: sheet names, row counts, last modified date.

        Returns:
            Dict with 'sheet_names', 'row_counts', 'file_size_bytes', 'last_modified',
            and optionally 'is_stale'.
        """
        self.validate_file(file_path)
        import openpyxl

        metadata: dict = {
            "file_size_bytes": os.path.getsize(file_path),
        }

        # Last modified time from OS
        mtime = os.path.getmtime(file_path)
        last_modified = datetime.fromtimestamp(mtime)
        metadata["last_modified"] = last_modified.isoformat()
        age_days = (datetime.now() - last_modified).days
        metadata["age_days"] = age_days
        if age_days > STALE_THRESHOLD_DAYS:
            metadata["is_stale"] = True
            logger.warning(
                "Excel file %s is %d days old (stale threshold: %d days).",
                file_path,
                age_days,
                STALE_THRESHOLD_DAYS,
            )
        else:
            metadata["is_stale"] = False

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            metadata["sheet_names"] = wb.sheetnames
            row_counts: dict[str, int] = {}
            for name in wb.sheetnames:
                ws = wb[name]
                row_counts[name] = ws.max_row if ws.max_row else 0
            metadata["row_counts"] = row_counts
            wb.close()
        except Exception:
            logger.warning("Failed to read workbook metadata from %s", file_path, exc_info=True)

        return metadata

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _read_all_rows_openpyxl(
        file_path: str,
        sheet_name: str | None = None,
    ) -> list[list]:
        """Read all rows from an Excel sheet using openpyxl."""
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            if ws is None:
                return []
            rows: list[list] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            return rows
        finally:
            wb.close()

    @staticmethod
    def _detect_header(rows: list[list]) -> tuple[int | None, list]:
        """Find the first row that looks like a header (contains >= 2 non-empty text cells).

        Returns:
            (header_index, header_cells) or (None, []) if not found.
        """
        for idx, row in enumerate(rows):
            text_cells = [cell for cell in row if cell is not None and isinstance(cell, str) and cell.strip()]
            if len(text_cells) >= 2:
                return idx, row
        return None, []

    @staticmethod
    def _find_name_value_cols(
        headers: list,
        data_rows: list[list],
    ) -> tuple[int | None, int | None]:
        """Identify which column holds the name (text) and which holds the numeric value.

        Uses a simple heuristic: first column that is predominantly text = name,
        first column that is predominantly numeric = value.
        """
        if not data_rows or not headers:
            return None, None

        num_cols = len(headers)
        text_counts = [0] * num_cols
        num_counts = [0] * num_cols

        sample = data_rows[:20]  # check first 20 rows
        for row in sample:
            for ci in range(min(len(row), num_cols)):
                cell = row[ci]
                if cell is None:
                    continue
                if isinstance(cell, (int, float)):
                    num_counts[ci] += 1
                elif isinstance(cell, str) and cell.strip():
                    # Could be a number formatted as string
                    try:
                        float(cell.replace(",", ""))
                        num_counts[ci] += 1
                    except (ValueError, AttributeError):
                        text_counts[ci] += 1

        name_col: int | None = None
        value_col: int | None = None

        # Pick the first predominantly-text column as name
        for ci in range(num_cols):
            if text_counts[ci] > num_counts[ci] and text_counts[ci] > 0:
                name_col = ci
                break

        # Pick the first predominantly-numeric column as value
        for ci in range(num_cols):
            if num_counts[ci] > text_counts[ci] and num_counts[ci] > 0:
                value_col = ci
                break

        return name_col, value_col

    @staticmethod
    def _to_float(value) -> float | None:
        """Attempt to convert a cell value to float."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.strip().replace(",", "").replace("₪", "").strip()
            if not cleaned:
                return None
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None

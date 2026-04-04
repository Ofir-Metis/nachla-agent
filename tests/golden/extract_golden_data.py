"""Extract golden test data from 25 example calculation Excel files.

Reads each Excel file and extracts:
- Settlement name and shovi value
- Nachla equivalent sqm values
- Hivun 3.75% and 33% amounts
- Split (pitzul) costs
- Per-building hasdarah data (where available)
- Priority area classification

Outputs a JSON file with all test cases.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# Base directory for golden test data
GOLDEN_DIR = Path(__file__).parent / "בדיקות התכנות"
OUTPUT_PATH = Path(__file__).parent / "golden_cases.json"


def _safe_float(value) -> float | None:
    """Safely convert a cell value to float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("₪", "").strip()
        if not cleaned or cleaned == "-" or cleaned.startswith("#"):
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _find_xlsx_file(folder_path: Path) -> Path | None:
    """Find the single xlsx file in a folder (the calculation file)."""
    xlsx_files = list(folder_path.glob("*.xlsx"))
    if not xlsx_files:
        return None
    # Filter out temp files
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]
    if len(xlsx_files) == 1:
        return xlsx_files[0]
    # If multiple, prefer one with 'תחשיבים' in name
    for f in xlsx_files:
        if "תחשיבים" in f.name or "תחשיב" in f.name:
            return f
    return xlsx_files[0]


def _extract_rate_sheet(wb: openpyxl.Workbook) -> dict | None:
    """Extract data from the main rate sheet ('תחשיב לפי שווי שומה (2)')."""
    sheet_name = "תחשיב לפי שווי שומה (2)"
    if sheet_name not in wb.sheetnames:
        logger.warning("Rate sheet not found")
        return None

    ws = wb[sheet_name]
    data = {}

    # Read all cells into a dict for flexible lookup
    cells = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 50, values_only=False):
        for cell in row:
            if cell.value is not None:
                cells[cell.coordinate] = cell.value

    # Find key values by scanning labels in column A.
    # Only scan first 30 rows -- the main rate table is always at the top.
    # Later rows may contain secondary calculations that would overwrite
    # primary values (e.g., Lobel file has a second 'היוון 3.75%' at row 59).
    scan_limit = min((ws.max_row or 50), 30)
    for row_num in range(1, scan_limit + 1):
        label = cells.get(f"A{row_num}")
        if label is None:
            continue
        label_str = str(label).strip()

        b_val = _safe_float(cells.get(f"B{row_num}"))
        c_val = cells.get(f"C{row_num}")
        d_val = _safe_float(cells.get(f"D{row_num}"))
        f_val = _safe_float(cells.get(f"F{row_num}"))
        h_val = _safe_float(cells.get(f"H{row_num}"))

        if "נחלה היוון מלא 3 דונם" in label_str:
            data.setdefault("sqm_nachla_3dunam", b_val)
        elif "נחלה היוון מלא 2.5 דונם" in label_str:
            data.setdefault("sqm_nachla_25dunam", b_val)
        elif ("פוטנציאל" in label_str and "אקו" in label_str) or label_str == "מ\"ר אקו' פוטנציאל":
            data.setdefault("sqm_potential", b_val)
        elif "דמי היתר שנרכשו" in label_str:
            data.setdefault("prior_permit_fees_sqm", b_val)
        elif "נחלה 3.75%" in label_str:
            data.setdefault("sqm_375", b_val)
        elif "מגרש מפוצל" in label_str:
            data.setdefault("sqm_split_plot", b_val)
            if d_val is not None:
                data.setdefault("sqm_split_discounted", d_val)
        elif "שווי מ\"ר אקו' מגורים" in label_str or ('שווי מ"ר' in label_str and "מגורים" in label_str):
            data.setdefault("shovi_per_sqm", b_val)
            # Also check C/D for office rates
            c_label = cells.get(f"C{row_num}")
            if c_label and "משרדים" in str(c_label):
                data.setdefault("shovi_offices", d_val)
        elif "היוון 3.75%" in label_str:
            data.setdefault("hivun_375_amount", f_val)
            data.setdefault("hivun_375_net", h_val)
            data.setdefault("hivun_375_rate", _safe_float(cells.get(f"C{row_num}")))
        elif "פיצול - חלק בהנחה" in label_str:
            data.setdefault("split_discounted_amount", f_val)
            data.setdefault("split_discounted_net", h_val)
        elif "פיצול חלק ללא הנחה" in label_str:
            data.setdefault("split_undiscounted_amount", f_val)
            data.setdefault("split_undiscounted_net", h_val)
        elif label_str.startswith('סה"כ פיצול'):
            data.setdefault("split_total_amount", f_val)
            data.setdefault("split_total_net", h_val)
        elif "היטל השבחה על פיצול" in label_str:
            data.setdefault("betterment_levy_split", f_val)
        elif label_str.startswith("דמי רכישה"):
            c_rate = _safe_float(c_val)
            data.setdefault("purchase_total_sqm", b_val)
            if "20%" in label_str or (c_rate is not None and abs(c_rate - 0.2) < 0.01):
                data.setdefault("purchase_20_amount", f_val)
                data.setdefault("purchase_20_net", h_val)
                data.setdefault("purchase_rate", c_rate)
            else:
                data.setdefault("purchase_33_amount", f_val)
                data.setdefault("purchase_33_net", h_val)
                data.setdefault("purchase_rate", c_rate)
        elif 'דמי היתר 1 מ"ר מגורים' in label_str:
            data.setdefault("permit_fee_per_sqm_main", f_val)
        elif 'דמי היתר 1 מ"ר שירות' in label_str:
            data.setdefault("permit_fee_per_sqm_service", f_val)
        elif 'דמי היתר 1 מ"ר מרתף' in label_str:
            data.setdefault("permit_fee_per_sqm_basement", f_val)
        elif 'דמי שימוש 1 מ"ר למגורים' in label_str:
            data.setdefault("usage_fee_per_sqm_residential", f_val)
        elif "דמי שימוש משרדים" in label_str:
            data.setdefault("usage_fee_per_sqm_office", f_val)
        elif "דמי היתר בריכה" in label_str:
            data.setdefault("permit_fee_per_sqm_pool", f_val)
        elif "דמי היתר משרד פלח" in label_str:
            data.setdefault("permit_fee_per_sqm_plach", f_val)

    return data


def _extract_sqm_sheet(wb: openpyxl.Workbook) -> dict | None:
    """Extract data from the sqm equivalent sheet."""
    sheet_name = "מ\"ר אקו' נחלה"
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    data = {}

    # Scan for key values
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 40), values_only=False):
        cells_in_row = {c.column: c.value for c in row if c.value is not None}
        label = cells_in_row.get(1)  # Column A
        if label is None:
            continue
        label_str = str(label).strip()
        b_val = _safe_float(cells_in_row.get(2))
        d_val = _safe_float(cells_in_row.get(4))

        if "גודל חלקת מגורים" in label_str:
            data["plot_size_sqm"] = b_val
        elif label_str == "עיקרי":
            if b_val is not None and "main_sqm" not in data:
                data["main_sqm"] = b_val
        elif "מרחבים מוגנים" in label_str or label_str == 'ממ"ד':
            if b_val is not None and "mamad_sqm" not in data:
                data["mamad_sqm"] = b_val
        elif 'יתרת ש"ש' in label_str:
            if b_val is not None and "service_sqm" not in data:
                data["service_sqm"] = b_val
        elif 'סה"כ תכסית' in label_str:
            if b_val is not None:
                data["building_coverage_sqm"] = b_val
        elif 'סה"כ מ"ר אקו\'' in label_str and "מעוגל" not in label_str:
            if d_val is not None and "total_sqm_equivalent" not in data:
                data["total_sqm_equivalent"] = d_val
        elif 'סה"כ מ"ר אקו\' מעוגל' in label_str:
            if d_val is not None and "total_sqm_rounded" not in data:
                data["total_sqm_rounded"] = d_val
        elif "עלויות פיתוח" in label_str and d_val is not None and "development_costs" not in data:
            data["development_costs"] = d_val

    return data


def _extract_hasdarah_sheet(wb: openpyxl.Workbook) -> list[dict] | None:
    """Extract per-building calculation data from hasdarah sheet."""
    sheet_name = "תחשיבי הסדרה"
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    buildings = []

    # Look for structured building data
    # The sheet has variable formats. Try to find header row first.
    header_row = None
    for row_num in range(1, min((ws.max_row or 0) + 1, 5)):
        for col_num in range(1, min((ws.max_column or 0) + 1, 12)):
            cell_val = ws.cell(row=row_num, column=col_num).value
            if cell_val and "מס' מבנה" in str(cell_val):
                header_row = row_num
                break
        if header_row:
            break

    if header_row is None:
        return None

    # Map header columns
    headers = {}
    for col_num in range(1, min((ws.max_column or 0) + 1, 12)):
        val = ws.cell(row=header_row, column=col_num).value
        if val:
            headers[str(val).strip()] = col_num

    # Read building rows
    for row_num in range(header_row + 1, min((ws.max_row or 0) + 1, 30)):
        building_id = ws.cell(row=row_num, column=headers.get("מס' מבנה", 1)).value
        if building_id is None:
            continue

        building = {"id": str(building_id)}

        for key, col in headers.items():
            val = ws.cell(row=row_num, column=col).value
            if val is not None:
                building[key] = val

        # Extract specific numeric fields
        for field_name, header_key in [
            ("permit_fees", "דמי היתר"),
            ("betterment_levy", "היטל השבחה"),
            ("total_cost", 'סה"כ למבנה'),
            ("usage_fees", "דמי שימוש"),
            ("area_sqm", "גודל"),
            ("total", 'סה"כ'),
        ]:
            for h_key, h_col in headers.items():
                if header_key in h_key:
                    val = _safe_float(ws.cell(row=row_num, column=h_col).value)
                    if val is not None:
                        building[field_name] = val

        name_col = headers.get("שם מבנה") or headers.get("סיווג המבנה")
        if name_col:
            building["name"] = str(ws.cell(row=row_num, column=name_col).value or "")

        action_col = headers.get("פעולה")
        if action_col:
            building["action"] = str(ws.cell(row=row_num, column=action_col).value or "")

        buildings.append(building)

    # Also check for summary row
    for row_num in range(header_row + 1, min((ws.max_row or 0) + 1, 30)):
        for col_num in range(1, 5):
            cell_val = ws.cell(row=row_num, column=col_num).value
            if cell_val and 'סה"כ' in str(cell_val):
                summary = {"type": "summary"}
                for key, col in headers.items():
                    val = _safe_float(ws.cell(row=row_num, column=col).value)
                    if val is not None:
                        summary[key] = val
                buildings.append(summary)
                break

    return buildings if buildings else None


def _extract_priority_area(wb: openpyxl.Workbook) -> str | None:
    """Extract priority area classification from the control sheet."""
    sheet_name = "בקרה"
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 15), values_only=False):
        for cell in row:
            if cell.value and "עדיפות לאומית" in str(cell.value):
                # Check adjacent cells for the classification
                row_num = cell.row
                for col in range(cell.column + 1, cell.column + 5):
                    val = ws.cell(row=row_num, column=col).value
                    if val:
                        val_str = str(val).strip()
                        if val_str in ("א", "A", "א'"):
                            return "A"
                        elif val_str in ("ב", "B", "ב'"):
                            return "B"
                        elif "עימות" in val_str or "frontline" in val_str.lower():
                            return "frontline"
                # Check next row for classification
                row_num2 = row_num + 1
                for col in range(1, 10):
                    val = ws.cell(row=row_num2, column=col).value
                    if val:
                        val_str = str(val).strip()
                        if val_str in ("א", "A", "א'"):
                            return "A"
                        elif val_str in ("ב", "B", "ב'"):
                            return "B"
                        elif "עימות" in val_str:
                            return "frontline"

    return None  # Standard (no priority area)


def _extract_settlement_name(folder_name: str) -> str:
    """Extract settlement name from folder name (after dash)."""
    parts = folder_name.split(" - ")
    if len(parts) >= 2:
        # Settlement name is usually the second part, remove annotations
        settlement = parts[1].strip()
        # Remove common suffixes
        for suffix in [
            "בית שלישי מאושר בתבע",
            "משק מהוון",
            "משק מורכב הרבה מבנים",
        ]:
            settlement = settlement.replace(suffix, "").strip()
        # If there's a third dash, settlement is just the second part
        return settlement.strip(" -")
    return folder_name


def extract_case(xlsx_path: Path, folder_name: str) -> dict | None:
    """Extract one golden test case from an Excel file.

    Returns dict with:
    - case_name: str (folder name)
    - source_file: str (xlsx filename)
    - inputs: extracted input values
    - expected: extracted expected result values
    """
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except Exception as e:
        logger.error("Failed to open %s: %s", xlsx_path.name, e)
        return None

    try:
        rate_data = _extract_rate_sheet(wb)
        sqm_data = _extract_sqm_sheet(wb)
        hasdarah_data = _extract_hasdarah_sheet(wb)
        priority_area = _extract_priority_area(wb)
        settlement_name = _extract_settlement_name(folder_name)

        if rate_data is None:
            logger.warning("No rate data found in %s", xlsx_path.name)
            return None

        shovi = rate_data.get("shovi_per_sqm")
        if shovi is None or shovi == 0:
            logger.warning("No shovi found in %s", xlsx_path.name)
            return None

        # Skip template files
        if folder_name == "טמפלט":
            logger.info("Skipping template folder")
            return None

        # Build inputs
        inputs = {
            "settlement_name": settlement_name,
            "shovi_per_sqm": shovi,
            "priority_area": priority_area,
        }

        if sqm_data:
            inputs["plot_size_sqm"] = sqm_data.get("plot_size_sqm", 2500)
            inputs["main_sqm"] = sqm_data.get("main_sqm")
            inputs["mamad_sqm"] = sqm_data.get("mamad_sqm")
            inputs["service_sqm"] = sqm_data.get("service_sqm")
            inputs["building_coverage_sqm"] = sqm_data.get("building_coverage_sqm")
            inputs["development_costs"] = sqm_data.get("development_costs", 0)
        else:
            inputs["plot_size_sqm"] = 2500

        inputs["prior_permit_fees_sqm"] = rate_data.get("prior_permit_fees_sqm", 0)

        # Determine if capitalized from folder name
        inputs["is_capitalized"] = "מהוון" in folder_name

        # Build expected values
        expected = {}

        # Sqm equivalent values
        if rate_data.get("sqm_nachla_25dunam") is not None:
            expected["sqm_nachla_25dunam"] = rate_data["sqm_nachla_25dunam"]
        if rate_data.get("sqm_nachla_3dunam") is not None:
            expected["sqm_nachla_3dunam"] = rate_data["sqm_nachla_3dunam"]
        if rate_data.get("sqm_375") is not None:
            expected["sqm_375"] = rate_data["sqm_375"]
        if rate_data.get("sqm_potential") is not None:
            expected["sqm_potential"] = rate_data["sqm_potential"]

        # Hivun 3.75%
        if rate_data.get("hivun_375_amount") is not None:
            expected["hivun_375_gross"] = rate_data["hivun_375_amount"]
        if rate_data.get("hivun_375_net") is not None:
            expected["hivun_375_net"] = rate_data["hivun_375_net"]

        # Hivun 33% (purchase)
        if rate_data.get("purchase_33_amount") is not None:
            expected["purchase_33_gross"] = rate_data["purchase_33_amount"]
        if rate_data.get("purchase_33_net") is not None:
            expected["purchase_33_net"] = rate_data["purchase_33_net"]

        # Hivun 20% (if exists)
        if rate_data.get("purchase_20_amount") is not None:
            expected["purchase_20_gross"] = rate_data["purchase_20_amount"]
        if rate_data.get("purchase_20_net") is not None:
            expected["purchase_20_net"] = rate_data["purchase_20_net"]

        # Purchase rate (actual rate used -- may be 0.33 or 0.2)
        if rate_data.get("purchase_rate") is not None:
            expected["purchase_rate"] = rate_data["purchase_rate"]
        if rate_data.get("purchase_total_sqm") is not None:
            expected["purchase_total_sqm"] = rate_data["purchase_total_sqm"]

        # Split (pitzul) costs
        if rate_data.get("split_total_amount") is not None:
            expected["split_total_gross"] = rate_data["split_total_amount"]
        if rate_data.get("split_total_net") is not None:
            expected["split_total_net"] = rate_data["split_total_net"]
        if rate_data.get("split_discounted_amount") is not None:
            expected["split_discounted_gross"] = rate_data["split_discounted_amount"]
        if rate_data.get("split_undiscounted_amount") is not None:
            expected["split_undiscounted_gross"] = rate_data["split_undiscounted_amount"]

        # Split plot sqm
        if rate_data.get("sqm_split_plot") is not None:
            expected["sqm_split_plot"] = rate_data["sqm_split_plot"]
        if rate_data.get("sqm_split_discounted") is not None:
            expected["sqm_split_discounted"] = rate_data["sqm_split_discounted"]

        # Betterment levy on split
        if rate_data.get("betterment_levy_split") is not None:
            expected["betterment_levy_split"] = rate_data["betterment_levy_split"]

        # Per-sqm unit rates (for validating rate lookups)
        for key in [
            "permit_fee_per_sqm_main",
            "permit_fee_per_sqm_service",
            "permit_fee_per_sqm_basement",
            "permit_fee_per_sqm_pool",
            "permit_fee_per_sqm_plach",
            "usage_fee_per_sqm_residential",
            "usage_fee_per_sqm_office",
        ]:
            if rate_data.get(key) is not None:
                expected[key] = rate_data[key]

        # Per-building hasdarah data
        if hasdarah_data:
            expected["hasdarah_buildings"] = hasdarah_data

        # Check we have meaningful data
        if not expected:
            logger.warning("No expected values extracted from %s", xlsx_path.name)
            return None

        case = {
            "case_name": folder_name,
            "source_file": xlsx_path.name,
            "inputs": inputs,
            "expected": expected,
            "sheets_available": wb.sheetnames,
        }

        return case

    except Exception as e:
        logger.error("Failed to extract from %s: %s", xlsx_path.name, e)
        return None
    finally:
        wb.close()


def extract_all_cases(golden_dir: Path) -> list[dict]:
    """Extract all golden test cases from the directory."""
    cases = []
    skipped = []

    if not golden_dir.exists():
        logger.error("Golden directory not found: %s", golden_dir)
        return []

    folders = sorted(golden_dir.iterdir())
    for folder in folders:
        if not folder.is_dir():
            continue

        xlsx_path = _find_xlsx_file(folder)
        if xlsx_path is None:
            logger.warning("No xlsx file found in %s", folder.name)
            skipped.append(folder.name)
            continue

        logger.info("Processing: %s -> %s", folder.name, xlsx_path.name)
        case = extract_case(xlsx_path, folder.name)

        if case is not None:
            cases.append(case)
            logger.info(
                "  Extracted: shovi=%s, hivun375=%s, hivun33=%s",
                case["inputs"].get("shovi_per_sqm"),
                case["expected"].get("hivun_375_gross"),
                case["expected"].get("purchase_33_gross"),
            )
        else:
            skipped.append(folder.name)

    logger.info(
        "\nExtraction complete: %d cases extracted, %d skipped",
        len(cases),
        len(skipped),
    )
    if skipped:
        logger.info("Skipped: %s", ", ".join(skipped))

    return cases


def main():
    """Run extraction and save to golden_cases.json."""
    logger.info("Starting golden data extraction from %s", GOLDEN_DIR)
    cases = extract_all_cases(GOLDEN_DIR)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(cases, f, ensure_ascii=False, indent=2, default=str)

    logger.info("Saved %d cases to %s", len(cases), OUTPUT_PATH)

    # Print summary
    print("\n" + "=" * 60)
    print(f"Golden test data: {len(cases)} cases extracted")
    print("=" * 60)
    for case in cases:
        inp = case["inputs"]
        exp = case["expected"]
        print(
            f"  {case['case_name']}: "
            f"shovi={inp.get('shovi_per_sqm')}, "
            f"hivun375={exp.get('hivun_375_gross', 'N/A')}, "
            f"hivun33={exp.get('purchase_33_gross', 'N/A')}"
        )


if __name__ == "__main__":
    main()
